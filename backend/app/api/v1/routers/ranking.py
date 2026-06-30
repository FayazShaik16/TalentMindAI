from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Any, Dict, List

from app.api.v1.dependencies.db import get_db
from app.api.v1.dependencies.auth import get_current_user
from app.schemas.responses import EnvelopeResponse
from app.schemas.ranking import (
    RankingRunRequest, CandidateRankingResponse, RankingResponse,
    RecommendationRequest
)
from app.database.repositories.job import JobRepository
from app.database.repositories.candidate import CandidateRepository
from app.database.repositories.ranking import RankingRepository
from app.database.models.ranking import JobCandidateRanking
from app.api.v1.routers.candidate_intelligence import sanitize_for_json
from app.services.agents.orchestrator import orchestrator
from app.core.logging.logging import logger
from app.core.config.config import settings

router = APIRouter(tags=["Hybrid Matching & Intelligent Ranking Engine"])

def map_db_to_response(db_rank: JobCandidateRanking) -> RankingResponse:
    """
    Maps SQLAlchemy model to Pydantic schema.
    """
    return RankingResponse(
        job_id=db_rank.job_id,
        rankings=[CandidateRankingResponse(**r) for r in db_rank.rankings],
        trace=db_rank.trace,
        statistics=db_rank.statistics
    )

@router.post("/ranking/run", response_model=EnvelopeResponse[RankingResponse])
async def run_ranking(
    payload: RankingRunRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Runs first-stage scoring + Cross-Encoder reranking for candidates against a Job ID,
    persisting results to the database.
    """
    logger.info("api_run_ranking_start", job_id=payload.job_id)

    # 1. Verify job exists
    job_repo = JobRepository(db)
    job = await job_repo.get_by_id(payload.job_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job Description with ID '{payload.job_id}' not found."
        )

    # 2. Run Hybrid Ranking Agent pipeline
    from app.services.agents.ranking_agent import ranking_agent
    await ranking_agent.initialize()

    context = {
        "db": db,
        "candidate_ids": payload.candidate_ids,
        "weights": payload.weights,
        "top_k_rerank": payload.top_k_rerank if payload.top_k_rerank is not None else settings.TOP_K_RERANK
    }

    try:
        final_output, _, _ = await orchestrator.execute_pipeline(
            pipeline=["hybrid_ranking"],
            initial_input=payload.job_id,
            context=context
        )
    except Exception as e:
        logger.exception("api_run_ranking_pipeline_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ranking pipeline execution failed: {str(e)}"
        )

    # 3. Retrieve the newly created db ranking record
    ranking_repo = RankingRepository(db)
    db_ranking = await ranking_repo.get_ranking(payload.job_id)
    if not db_ranking:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve ranking from database after execution."
        )

    await db.commit()

    return EnvelopeResponse(data=map_db_to_response(db_ranking))

@router.post("/ranking/rebuild", response_model=EnvelopeResponse[RankingResponse])
async def rebuild_ranking(
    payload: RankingRunRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Rebuilds rankings for a Job ID using updated weights or candidates list.
    """
    return await run_ranking(payload, db, current_user)

@router.get("/ranking/{job_id}", response_model=EnvelopeResponse[RankingResponse])
async def get_ranking(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Get full candidate ranking details stored for a Job ID.
    """
    repo = RankingRepository(db)
    ranking = await repo.get_ranking(job_id)
    if not ranking:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Ranking records for Job ID '{job_id}' not found. Trigger ranking run first."
        )
    return EnvelopeResponse(data=map_db_to_response(ranking))

@router.get("/ranking/{job_id}/top", response_model=EnvelopeResponse[List[CandidateRankingResponse]])
async def get_top_rankings(
    job_id: str,
    limit: int = 10,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Returns only the top K ranked candidates for a Job ID.
    """
    repo = RankingRepository(db)
    ranking = await repo.get_ranking(job_id)
    if not ranking:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Ranking records for Job ID '{job_id}' not found."
        )
    
    mapped = map_db_to_response(ranking)
    top_list = mapped.rankings[:limit]
    return EnvelopeResponse(data=top_list)

@router.get("/ranking/{job_id}/trace", response_model=EnvelopeResponse[List[Dict[str, Any]]])
async def get_ranking_trace(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Get step-by-step process execution traces for a Job ID ranking run.
    """
    repo = RankingRepository(db)
    ranking = await repo.get_ranking(job_id)
    if not ranking:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Ranking records for Job ID '{job_id}' not found."
        )
    return EnvelopeResponse(data=ranking.trace)

@router.get("/ranking/{job_id}/statistics", response_model=EnvelopeResponse[Dict[str, Any]])
async def get_ranking_statistics(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Get benchmarking CPU, memory and processing duration statistics for a Job ID ranking run.
    """
    repo = RankingRepository(db)
    ranking = await repo.get_ranking(job_id)
    if not ranking:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Ranking records for Job ID '{job_id}' not found."
        )
    return EnvelopeResponse(data=ranking.statistics)

@router.post("/recommendations", response_model=EnvelopeResponse[Dict[str, Any]])
async def get_single_recommendation(
    payload: RecommendationRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Quick candidate-job recommendation analysis for a single candidate.
    """
    # Verify candidate and job exist
    cand_repo = CandidateRepository(db)
    cand = await cand_repo.get_candidate_profile(payload.candidate_id)
    if not cand:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Candidate with ID '{payload.candidate_id}' not found."
        )

    job_repo = JobRepository(db)
    job = await job_repo.get_by_id(payload.job_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job Description with ID '{payload.job_id}' not found."
        )

    # Trigger a ranking run focusing purely on this candidate
    from app.services.agents.ranking_agent import ranking_agent
    await ranking_agent.initialize()

    context = {
        "db": db,
        "candidate_ids": [payload.candidate_id],
        "top_k_rerank": 0  # no need to rerank since there is only one candidate
    }

    try:
        final_output, _, _ = await orchestrator.execute_pipeline(
            pipeline=["hybrid_ranking"],
            initial_input=payload.job_id,
            context=context
        )
        
        rankings = final_output.get("rankings", [])
        if not rankings:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Scoring pipeline returned empty results."
            )
        
        cand_data = rankings[0]
        return EnvelopeResponse(data={
            "job_id": payload.job_id,
            "candidate_id": payload.candidate_id,
            "overall_score": cand_data["overall_score"],
            "hiring_confidence": cand_data["hiring_confidence"],
            "recommendation": cand_data["recommendation"],
            "reasoning_summary": cand_data["reasoning_summary"],
            "interview_recommendation": cand_data["interview_recommendation"],
            "missing_skills": cand_data["missing_skills"]
        })
    except Exception as e:
        logger.exception("api_recommendation_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate recommendation: {str(e)}"
        )

@router.get("/ranking/{job_id}/export-csv")
async def export_ranking_csv(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Export ranking results to a CSV file in-memory.
    """
    repo = RankingRepository(db)
    ranking = await repo.get_ranking(job_id)
    if not ranking or not ranking.rankings:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Ranking records for Job ID '{job_id}' not found. Trigger ranking run first."
        )

    import io
    import csv
    from fastapi import Response
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        "Rank", "Candidate ID", "First Name", "Last Name", 
        "Overall Score", "Hiring Confidence", "Recommendation", 
        "Growth Potential", "Missing Skills", "Reasoning Summary", 
        "Evidence Summary", "Risk Summary"
    ]
    
    writer.writerow(headers)
    for r in ranking.rankings:
        missing_skills_str = ", ".join(r.get("missing_skills", []))
        writer.writerow([
            r.get("rank"),
            r.get("candidate_id"),
            r.get("first_name"),
            r.get("last_name"),
            r.get("overall_score"),
            r.get("hiring_confidence"),
            r.get("recommendation"),
            r.get("growth_potential"),
            missing_skills_str,
            r.get("reasoning_summary"),
            r.get("evidence_summary"),
            r.get("risk_summary")
        ])
            
    content = output.getvalue()
    output.close()
    
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=rankings_{job_id}.csv"}
    )

@router.get("/ranking/{job_id}/export-xlsx")
async def export_ranking_xlsx(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Export ranking results to a beautifully formatted Excel sheet in-memory.
    """
    repo = RankingRepository(db)
    ranking = await repo.get_ranking(job_id)
    if not ranking or not ranking.rankings:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Ranking records for Job ID '{job_id}' not found. Trigger ranking run first."
        )

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Match Rankings"
    
    # Title Block
    ws.merge_cells("A1:L1")
    ws["A1"] = f"TalentMind AI - Candidate Match Rankings for Job ID: {job_id}"
    ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = [
        "Rank", "Candidate ID", "First Name", "Last Name", 
        "Overall Score", "Hiring Confidence", "Recommendation", 
        "Growth Potential", "Missing Skills", "Reasoning Summary", 
        "Evidence Summary", "Risk Summary"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25
    
    # Border styles
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Add rankings data
    for row_num, r in enumerate(ranking.rankings, 3):
        ws.row_dimensions[row_num].height = 20
        row_data = [
            r.get("rank"),
            r.get("candidate_id"),
            r.get("first_name"),
            r.get("last_name"),
            r.get("overall_score") / 100.0 if r.get("overall_score") is not None else None,
            r.get("hiring_confidence"),
            r.get("recommendation"),
            r.get("growth_potential") / 100.0 if r.get("growth_potential") is not None else None,
            ", ".join(r.get("missing_skills", [])),
            r.get("reasoning_summary"),
            r.get("evidence_summary"),
            r.get("risk_summary")
        ]
        
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = border
            
            # Alignments & formatting
            if col_num in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [5, 6, 8]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.0%'
            elif col_num == 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Highlight recommendation colors
                rec_val = str(val).lower()
                if "hire" in rec_val or "recommend" in rec_val:
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, color="15803D", bold=True)
                elif "disqualify" in rec_val or "reject" in rec_val:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, color="B91C1C", bold=True)
                else:
                    cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=10, color="A16207", bold=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Adjust column widths automatically
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        if col[0].column in [9, 10, 11, 12]:
            ws.column_dimensions[col_letter].width = 30
            continue
            
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                if isinstance(cell.value, float) and cell.number_format == '0.0%':
                    val_str = f"{cell.value * 100:.1f}%"
                else:
                    val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    content = output.getvalue()
    output.close()
    
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rankings_{job_id}.xlsx"}
    )

